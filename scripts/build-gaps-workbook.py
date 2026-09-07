# Builds the Excel workbook Steve and Gaetana actually work in.
#
#   npm run report:cartons && npm run report:bulky   (refresh the source data)
#   python3 scripts/build-gaps-workbook.py           -> reports/MasterKraft-Product-Data-Gaps.xlsx
#
# The markdown reports are for reading. This is for DOING: the two fill-in sheets
# have yellow input cells that go straight back into WordPress, and the summary
# counts with formulas so it stays honest as rows are worked through.
#
# Source of truth is the CSVs, never this file. Re-run the reports first.
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPORTS = Path("reports")
OUT = REPORTS / "MasterKraft-Product-Data-Gaps.xlsx"

FONT = "Arial"
INK = "15181B"
ACCENT = "B93C26"
MUTED = "5B656D"

HEAD_FILL = PatternFill("solid", fgColor="15181B")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")   # fill these in
NOTE_FILL = PatternFill("solid", fgColor="F0F3F4")
THIN = Side(style="thin", color="C9D1D4")
BORDER = Border(bottom=THIN)


def read(name):
    with open(REPORTS / name, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def title(ws, text, subtitle):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color=MUTED)
    ws.row_dimensions[1].height = 22


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def body(ws, row, values, input_cols=(), italic=False, fill=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10, italic=italic,
                      color=MUTED if italic else INK)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=(i == 3))
        if i in input_cols:
            c.fill = fill or INPUT_FILL


def legend(ws, row, lines):
    # Column B, never A. Column A is counted by the summary formulas, so a line
    # of prose in it would be counted as a product.
    for n, line in enumerate(lines):
        c = ws.cell(row=row + n, column=2, value=line)
        c.font = Font(name=FONT, size=9, color=MUTED, bold=(n == 0))
        if n == 0:
            c.font = Font(name=FONT, size=9, bold=True, color=ACCENT)
    return row + len(lines) + 1


gaps = read("carton-gaps.csv")
bulky = [r for r in read("bulky-freight-profile.csv") if r["segment"] == "bulky"]

wb = Workbook()

# ------------------------------------------------------------------ summary
ws = wb.active
ws.title = "Summary"
title(ws, "Product data gaps: what needs doing",
      "Generated from reports/carton-gaps.csv. Counts are formulas, so they follow the sheets.")

header_row(ws, 4, ["Pile", "Products", "Who does it", "Effort"], [34, 11, 30, 40])
n_units = sum(1 for g in gaps if g["pile"] == "1-units")
n_sets = sum(1 for g in gaps if g["pile"] == "2-set")
n_measure = sum(1 for g in gaps if g["pile"] == "3-measure")

rows = [
    ("1. Wrong units, live on the site now", "'1 Fix units'", n_units, "Whoever edits WordPress", "Minutes. Do these first."),
    ("2. Sets and multi-packs", "'2 Sets decision'", n_sets, "A commercial decision", "A decision, not a task."),
    ("3. Genuinely unmeasured", "'3 Measure'", n_measure, "Someone with a tape measure", "The only physical work."),
]
for n, (label, sheet, count, who, effort) in enumerate(rows):
    r = 5 + n
    body(ws, r, [label, None, who, effort])
    # COUNTA over exactly the data rows, not an open-ended range: row 5 is the
    # example and anything below the data is prose. Excel rewrites the range if
    # somebody deletes a completed row, so it stays right.
    last = 5 + count
    ws.cell(row=r, column=2, value=f"=COUNTA({sheet}!A6:A{last})")
    ws.cell(row=r, column=2).font = Font(name=FONT, size=10, color=INK)
    ws.cell(row=r, column=2).border = BORDER

r = 5 + len(rows)
ws.cell(row=r, column=1, value="Total blocked from online freight quoting").font = Font(name=FONT, size=10, bold=True, color=INK)
ws.cell(row=r, column=2, value="=SUM(B5:B7)").font = Font(name=FONT, size=10, bold=True, color=INK)

nxt = legend(ws, r + 2, [
    "How to use this",
    "Yellow cells are for you to fill in. Everything else is reference.",
    "Work sheet 1 first: those figures are wrong on the live website today, not merely missing.",
    "Sheet 2 needs a decision from the business before anyone touches data.",
    "Sheet 3 is the only pile that needs a physical tape measure.",
])
legend(ws, nxt, [
    "What this is not",
    "260812-Missing-ProductData.xlsx does not overlap with any of these (checked, zero matches).",
    "That file lists 73 products with no buy price, which is a margin problem, and 6 CBM rows which are volume only.",
])

# ------------------------------------------------- 1 fix units (fill-in sheet)
ws = wb.create_sheet("1 Fix units")
title(ws, "Pile 1: wrong units, live on the site now",
      "Recorded in millimetres inside a centimetre field. SCRWAR04 currently says its carton is 24 metres long.")
header_row(ws, 4, ["SKU", "Product", "Recorded now", "Suspected", "Correct L (cm)", "Correct W (cm)", "Correct H (cm)", "Correct kg"], [14, 40, 22, 26, 13, 13, 13, 11])

body(ws, 5, ["EXAMPLE", "This row is an example, delete it", "32.9kg L2440 W610", "suspected 244 x 61 x ?cm", 244, 61, 38, 32.9],
     input_cols=(5, 6, 7, 8), italic=True, fill=NOTE_FILL)

row = 6
for g in [x for x in gaps if x["pile"] == "1-units"]:
    rec = " ".join(p for p in [
        f'{g["recorded_kg"]}kg' if g["recorded_kg"] != "0" else "",
        f'L{g["recorded_l_cm"]}' if g["recorded_l_cm"] != "0" else "",
        f'W{g["recorded_w_cm"]}' if g["recorded_w_cm"] != "0" else "",
        f'H{g["recorded_h_cm"]}' if g["recorded_h_cm"] != "0" else "",
    ] if p) or "nothing"
    body(ws, row, [g["sku"], g["product"], rec, g["action"], None, None, None, None], input_cols=(5, 6, 7, 8))
    row += 1

legend(ws, row + 1, [
    "Notes",
    "Suspected values are the recorded figure divided by ten. Confirm ONE against a physical box before applying the rest, in case the error is not uniform.",
    "A '?' in the Suspected column means that dimension is missing entirely, so it needs measuring as well as correcting.",
    "Record the SHIPPING CARTON, not the assembled size of the equipment.",
])

# ------------------------------------------------------------ 2 sets decision
ws = wb.create_sheet("2 Sets decision")
title(ws, "Pile 2: sets and multi-packs",
      "A 150kg plate set has no single carton because it ships as several. No measurement resolves this.")
header_row(ws, 4, ["SKU", "Product", "Decision", "Notes"], [16, 46, 24, 46])
body(ws, 5, ["EXAMPLE", "This row is an example, delete it", "Manual quote", "Ships as 5 cartons, not worth a nominal figure"],
     input_cols=(3, 4), italic=True, fill=NOTE_FILL)
row = 6
for g in [x for x in gaps if x["pile"] == "2-set"]:
    body(ws, row, [g["sku"], g["product"], None, None], input_cols=(3, 4))
    row += 1

legend(ws, row + 1, [
    "The choice, per product",
    "Nominal consignment: give the set a combined carton figure built from its components, so the website can quote it.",
    "Manual quote: exclude it from online freight and send the customer to the quote flow.",
    "Manual quote is what happens today, by accident rather than by decision.",
])

# --------------------------------------------------- 3 measure (fill-in sheet)
ws = wb.create_sheet("3 Measure")
title(ws, "Pile 3: measure these",
      "The only pile needing a tape measure. Mostly small accessories, which would likely become parcel-shippable once measured.")
header_row(ws, 4, ["SKU", "Product", "Already known", "Carton L (cm)", "Carton W (cm)", "Carton H (cm)", "Gross kg", "Measured by"], [14, 40, 18, 13, 13, 13, 11, 16])
body(ws, 5, ["EXAMPLE", "This row is an example, delete it", "kg=11", 38, 30, 6, 11, "SW"],
     input_cols=(4, 5, 6, 7, 8), italic=True, fill=NOTE_FILL)
row = 6
for g in [x for x in gaps if x["pile"] == "3-measure"]:
    known = " ".join(p for p in [
        f'kg={g["recorded_kg"]}' if g["recorded_kg"] != "0" else "",
        f'L={g["recorded_l_cm"]}' if g["recorded_l_cm"] != "0" else "",
        f'W={g["recorded_w_cm"]}' if g["recorded_w_cm"] != "0" else "",
        f'H={g["recorded_h_cm"]}' if g["recorded_h_cm"] != "0" else "",
    ] if p) or "nothing"
    body(ws, row, [g["sku"], g["product"], known, None, None, None, None, None], input_cols=(4, 5, 6, 7, 8))
    row += 1

legend(ws, row + 1, [
    "How to measure",
    "The SHIPPING CARTON in centimetres, not the product's own size. Round up to the next whole centimetre.",
    "GROSS weight in kilograms: the product plus its packaging, because that is what freight is charged on.",
    "Where a weight is already known it is shown in 'Already known'. Only the missing figures need filling.",
])

# ------------------------------------------------------------ bulky reference
ws = wb.create_sheet("Bulky reference")
title(ws, "Reference: the 107 products no parcel network will carry",
      "Not a task list. Provided so a freight provider can price the real range. Over 22kg, over 105cm a side, or over 0.25 cubic metres.")
header_row(ws, 4, ["SKU", "Product", "kg", "L (cm)", "W (cm)", "H (cm)", "m3", "Outside parcel limits on"], [14, 46, 9, 10, 10, 10, 9, 26])
row = 5
for b in sorted(bulky, key=lambda x: -float(x["weight_kg"])):
    body(ws, row, [b["sku"], b["product"], float(b["weight_kg"]), float(b["length_cm"]),
                   float(b["width_cm"]), float(b["height_cm"]), float(b["cubic_m3"]),
                   b["out_of_parcel_because"]])
    row += 1
ws.auto_filter.ref = f"A4:H{row - 1}"

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"wrote {OUT}")
print(f"  1 Fix units      {sum(1 for g in gaps if g['pile'] == '1-units')}")
print(f"  2 Sets decision  {sum(1 for g in gaps if g['pile'] == '2-set')}")
print(f"  3 Measure        {sum(1 for g in gaps if g['pile'] == '3-measure')}")
print(f"  Bulky reference  {len(bulky)}")
